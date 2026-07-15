#calculando a média com openpyxl
import openpyxl

#ler arquivo excel
planilha = openpyxl.load_workbook("tarefa_python.xlsx")
#selecionar a planilha ativa
aba = planilha.active

#adiona os titulos nas novas colunas
aba["D1"] = "Média"
aba["E1"] = "Situação"

#calcular a média e atribuir a situação de aprovado ou reprovado
for linha in range(2, aba.max_row + 1):
    #float para garantir que as notas sejam tratadas como números decimais
    nota1 = float(aba.cell(row=linha, column=2).value)
    nota2 = float(aba.cell(row=linha, column=3).value)
    media = (nota1 + nota2) / 2
    aba.cell(row=linha, column=4).value = media
    
    #if para atribuir a situação de aprovado ou reprovado com base na média
    if media >= 6:
        aba.cell(row=linha, column=5).value = "Aprovado"
    else:
        aba.cell(row=linha, column=5).value = "Reprovado"
    #exibir resultado para cada aluno no terminal
    nome = aba.cell(row=linha, column=1).value
    situacao = aba.cell(row=linha, column=5).value
    
    #f-string para exibir o resultado formatado no terminal 
    print(f"Aluno: {nome} | Média: {media:.2f} | Situação: {situacao}")

#salvar o arquivo atualizado
planilha.save("resultado_openpyxl.xlsx")
print("Arquivo salvo com sucesso!")