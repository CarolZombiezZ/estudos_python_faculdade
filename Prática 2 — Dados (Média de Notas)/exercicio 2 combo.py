#combo de bibliotas
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Color

#calculos com pandas
tabela = pd.read_excel("tarefa_python.xlsx")
tabela["média"] = (tabela["Nota 1"] + tabela["Nota 2"]) / 2
tabela.loc[tabela["média"] >=6, "situação"] = "Aprovado"
tabela.loc[tabela["média"] <6, "situação"] = "Reprovado"

#salvando o resultado bruto
tabela.to_excel("resultado_final.xlsx", index=False)

#usando openpyxl para estetica
wb = load_workbook("resultado_final.xlsx")
ws = wb.active

#pintando texto de vermelho para reprovados e verde para aprovados
for linha in range(2, ws.max_row + 1):
    celula_situacao = ws.cell(row=linha, column=5)

    if celula_situacao.value == "Reprovado":
        celula_situacao.font = Font(color="FF0000", bold=True) # Vermelho e Negrito
    else:
        celula_situacao.font = Font(color="008000", bold=True) # Verde e Negrito

wb.save("resultado_colorido.xlsx")
print("Planilha processada pelo Pandas e formatada pelo Openpyxl!")